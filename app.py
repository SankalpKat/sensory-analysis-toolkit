import streamlit as st
import numpy as np
import pandas as pd
import os
from dotenv import load_dotenv
from supabase import create_client
import secrets

load_dotenv()

# Connect to Supabase — reads credentials from .env file
supabase = create_client(
    os.getenv("SUPABASE_URL"),
    os.getenv("SUPABASE_KEY")
)

st.set_page_config(
    page_title="Sensory Analysis Toolkit",
    page_icon="🧪",
    layout="wide"
)

st.title("Sensory Analysis Toolkit")
st.caption("Double-blind randomized sensory trial designer")

if "trial_saved" not in st.session_state:
    st.session_state["trial_saved"] = False

if "trial_config" not in st.session_state:
    st.session_state["trial_config"] = {}

def suggest_design(num_products, num_panelists, scale_type):
    reasons = []
    warnings = []

    if num_products > 6:
        suggested = "Balanced Incomplete Block (BIB)"
        reasons.append(f"You have {num_products} products. WLS becomes impractical above 6 — too many samples per panelist per session.")
        reasons.append("BIB lets each panelist evaluate a manageable subset while keeping all product pairs fairly compared.")
        if num_panelists < num_products * 2:
            warnings.append(f"With {num_products} products and {num_panelists} panelists, BIB may not fully balance. Consider more panelists.")
        return suggested, reasons, warnings

    if scale_type == "Preference (forced choice: pick one)":
        suggested = "Monadic Sequential"
        reasons.append("Forced-choice preference doesn't need crossover balancing — the comparison happens in the panelist's mind.")
        reasons.append("Randomizing serving order across panelists is sufficient to control position bias.")
        return suggested, reasons, warnings

    if scale_type == "100mm Visual Analogue Scale (VAS)":
        suggested = "Williams Latin Square"
        reasons.append("VAS is typically used with trained panels who evaluate all products. WLS is the standard crossover design for this.")
        if num_panelists % num_products != 0:
            req = num_products * round(num_panelists / num_products)
            warnings.append(f"WLS requires panelists to be a multiple of {num_products}. Consider adjusting to {req} panelists.")
        return suggested, reasons, warnings

    if num_products <= 6 and num_panelists >= num_products:
        suggested = "Williams Latin Square"
        reasons.append(f"With {num_products} products and {num_panelists} panelists, WLS is feasible and gives the strongest statistical control.")
        reasons.append("Every panelist evaluates every product, and serving order is fully counterbalanced.")
        if num_panelists % num_products != 0:
            req = num_products * round(num_panelists / num_products)
            warnings.append(f"WLS works best when panelists are a multiple of {num_products}. Consider adjusting to {req}.")
        return suggested, reasons, warnings

    suggested = "Monadic Sequential"
    reasons.append("Monadic Sequential is the most flexible design — works under any conditions.")
    return suggested, reasons, warnings

def generate_williams_latin_square(n):
    # Build the base interleaved sequence
    # Alternates taking from the low and high end: 0, n-1, 1, n-2, 2, n-3...
    first_row = []
    lo, hi = 0, n - 1
    turn = True
    while lo <= hi:
        if turn:
            first_row.append(lo)
            lo += 1
        else:
            first_row.append(hi)
            hi -= 1
        turn = not turn

    # Generate n rows by rotating the first row
    square = []
    for i in range(n):
        row = [first_row[(j + i) % n] for j in range(n)]
        square.append(row)

    # For odd n, Williams' method requires 2n rows for full balance
    # The second set is the mirror (reverse) of each row in the first set
    if n % 2 == 1:
        for i in range(n):
            square.append(list(reversed(square[i])))

    return square

def generate_monadic_sequential(num_products, num_panelists):
    # Shuffle product order independently for each panelist
    rows = []
    for _ in range(num_panelists):
        order = list(np.random.permutation(num_products))
        rows.append(order)
    return rows


def generate_bib(num_products, k, num_panelists):
    # All possible subsets of size k, cycled across panelists
    from itertools import combinations
    all_blocks = list(combinations(range(num_products), k))
    rows = []
    for i in range(num_panelists):
        block = all_blocks[i % len(all_blocks)]
        rows.append(list(block))
    return rows


# -----------------------------------------------------------------------------
# SIDEBAR
# -----------------------------------------------------------------------------
st.sidebar.title("⚙️ Trial Configuration")

trial_name = st.sidebar.text_input(label="Trial name", value="My Sensory Trial")

num_products = st.sidebar.number_input(
    label="Number of products", min_value=2, max_value=10, value=3, step=1
)

st.sidebar.markdown("**Product names**")
product_names = []
for i in range(num_products):
    name = st.sidebar.text_input(
        label=f"Product {i+1}",
        value=f"Product {chr(65+i)}",
        key=f"product_{i}"
    )
    product_names.append(name)

st.sidebar.markdown("---")

num_panelists = st.sidebar.number_input(
    label="Number of panelists", min_value=2, max_value=200, value=12, step=1
)

st.sidebar.markdown("---")

scale_type = st.sidebar.selectbox(
    label="Rating scale",
    options=[
        "9-point Hedonic",
        "100mm Visual Analogue Scale (VAS)",
        "5-point Likert",
        "Just About Right (JAR) — 5-point",
        "Preference (forced choice: pick one)"
    ]
)

st.sidebar.markdown("---")

suggested, reasons, warnings = suggest_design(num_products, num_panelists, scale_type)

with st.sidebar.expander("💡 Suggested design", expanded=True):
    st.markdown(f"**Recommendation: {suggested}**")
    for r in reasons:
        st.caption(r)
    for w in warnings:
        st.warning(w)
    st.caption("You can override this below.")

design_options = [
    "Williams Latin Square",
    "Balanced Incomplete Block (BIB)",
    "Monadic Sequential"
]

design_type = st.sidebar.selectbox(
    label="Trial design",
    options=design_options,
    index=design_options.index(suggested)
)

if design_type == "Balanced Incomplete Block (BIB)":
    samples_per_panelist = st.sidebar.number_input(
        label="Samples per panelist (k)",
        min_value=2,
        max_value=num_products - 1,
        value=min(3, num_products - 1),
        step=1
    )
else:
    samples_per_panelist = num_products

st.sidebar.markdown("---")

num_sessions = st.sidebar.number_input(
    label="Number of sessions", min_value=1, max_value=10, value=1, step=1
)

washout_days = st.sidebar.number_input(
    label="Washout period between sessions (days)", min_value=0, max_value=30, value=1, step=1
)

st.sidebar.markdown("---")
st.sidebar.markdown("---")
st.sidebar.markdown("**Survey questions**")
st.sidebar.caption("Overall liking is always included. Add more attributes below, one per line.")
attributes_input = st.sidebar.text_area(
    label="Additional attributes",
    value="Flavor\nTexture\nAppearance",
    height=120,
    help="Each line becomes a separate rating question on the ballot."
)
# Split by newline, strip whitespace, remove empty lines
# "Overall liking" is always first
custom_attributes = ["Overall liking"] + [
    a.strip() for a in attributes_input.split("\n") if a.strip()
]

st.sidebar.markdown("---")
save_clicked = st.sidebar.button(
    label="Save configuration", type="primary", use_container_width=True
)

if save_clicked:
    if len(set(product_names)) != len(product_names):
        st.sidebar.error("Product names must be unique.")
    else:
        st.session_state["trial_config"] = {
            "trial_name": trial_name,
            "num_products": num_products,
            "product_names": product_names,
            "num_panelists": num_panelists,
            "scale_type": scale_type,
            "design_type": design_type,
            "samples_per_panelist": samples_per_panelist,
            "num_sessions": num_sessions,
             "washout_days": washout_days,
            "attributes": custom_attributes,
        }
        st.session_state["trial_saved"] = True
        st.session_state["codes_locked"] = False
        st.session_state["blinding_codes"] = {}
        st.sidebar.success("Configuration saved.")


# -----------------------------------------------------------------------------
# MAIN PAGE
# -----------------------------------------------------------------------------
if not st.session_state["trial_saved"]:
    st.info("👈 Configure your trial in the sidebar to begin.")
    st.markdown("---")
    st.subheader("📖 Design Guide — Which method should you use?")
    st.markdown("Read below, or use the **💡 Suggested design** box in the sidebar — it updates as you change your inputs.")
    st.markdown("")

    st.markdown("**Quick comparison**")
    st.table({
        "": ["Products per panelist", "Panel size", "Carry-over controlled?", "Best for"],
        "Williams Latin Square": ["All products", "Multiple of product count", "Yes — explicitly balanced", "Trained panels, 2-6 products"],
        "Balanced Incomplete Block": ["A subset (k < total)", "Larger panels needed", "Partial — within each block", "7+ products, fatigue risk"],
        "Monadic Sequential": ["All products", "Any — 50-200+ typical", "No — relies on randomization", "Consumer acceptance testing"],
    })

    st.markdown("")

    with st.expander("🔲 Williams Latin Square"):
        st.markdown("**Every panelist evaluates every product, in a counterbalanced order.**")
        st.markdown(
            "A Latin Square arranges products so each appears exactly once in each serving position "
            "across the panel. Williams (1949) extended this to also balance carry-over: every product "
            "follows every other product the same number of times. Controls both position effects and carry-over simultaneously."
        )
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Best for**")
            st.markdown("- 2-6 products\n- Trained or semi-trained panels\n- Strong flavors where carry-over matters\n- Studies where statistical power is the priority")
        with col2:
            st.markdown("**Limitations**")
            st.markdown("- Panelist count must be a multiple of product count\n- Impractical above 6 products\n- May require multiple sessions")
        st.markdown("**Example** — 3 products, 6 panelists:")
        st.code("Panelist 1: A -> B -> C\nPanelist 2: B -> C -> A\nPanelist 3: C -> A -> B\nPanelist 4: A -> C -> B\nPanelist 5: C -> B -> A\nPanelist 6: B -> A -> C\n-> Each product appears once in each position", language=None)

    with st.expander("🧩 Balanced Incomplete Block (BIB)"):
        st.markdown("**Each panelist evaluates a subset — but every product pair is compared equally often.**")
        st.markdown(
            "When you have too many products for one panelist to evaluate in a session, BIB assigns "
            "each panelist a subset of size k. The design is balanced because every pair of products "
            "appears together in the same panelist's set the same number of times (lambda)."
        )
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Best for**")
            st.markdown("- 7+ products\n- Fatigue or time constraints\n- Screening studies with many candidates")
        with col2:
            st.markdown("**Limitations**")
            st.markdown("- Needs more panelists than WLS\n- Not all (products, k) combinations are valid\n- Less power per panelist than WLS")
        st.markdown("**Example** — 4 products, k=2, 6 panelists:")
        st.code("Panelist 1: A, B\nPanelist 2: A, C\nPanelist 3: A, D\nPanelist 4: B, C\nPanelist 5: B, D\nPanelist 6: C, D\n-> Every pair appears exactly once (lambda=1)", language=None)

    with st.expander("➡️ Monadic Sequential"):
        st.markdown("**Panelists evaluate products one at a time in a randomized order.**")
        st.markdown(
            "No systematic counterbalancing of carry-over — position bias is controlled by "
            "randomizing serving order independently for each panelist. Standard for large consumer "
            "panels where statistical control comes from panel size rather than design structure."
        )
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Best for**")
            st.markdown("- Consumer panels (naive panelists)\n- Hedonic and acceptance testing\n- Large panels (50-200+)\n- Forced-choice preference")
        with col2:
            st.markdown("**Limitations**")
            st.markdown("- Does not explicitly balance carry-over\n- Needs larger panel than WLS for same power\n- Not ideal for intense, long-lasting flavors")
        st.markdown("**Example** — 3 products, randomized per panelist:")
        st.code("Panelist 1: B -> A -> C\nPanelist 2: C -> B -> A\nPanelist 3: A -> C -> B\n-> Each order drawn independently at random", language=None)

else:
    config = st.session_state["trial_config"]
    st.success(
        f"Active trial: {config['trial_name']} — "
        f"{config['design_type']} — "
        f"{config['num_products']} products — "
        f"{config['num_panelists']} panelists"
    )

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📊 Design Matrix",
        "🔒 Blinding Codes",
        "📋 Ballots",
        "📈 Power Calculator",
        "📥 Data Entry Template",
        "🔗 Share Links"
    ])

    with tab1:
        st.subheader("Design Matrix")
        st.markdown(
            "This table shows which product each panelist evaluates in which position. "
            "Rows are panelists, columns are serving positions."
        )

        products = config["product_names"]
        n_products = config["num_products"]
        n_panelists = config["num_panelists"]
        design = config["design_type"]
        k = config["samples_per_panelist"]

        # Generate matrix based on chosen design
        if design == "Williams Latin Square":
            matrix = generate_williams_latin_square(n_products)
            full_matrix = []
            for i in range(n_panelists):
                row = matrix[i % len(matrix)]
                full_matrix.append(row)

        elif design == "Monadic Sequential":
            full_matrix = generate_monadic_sequential(n_products, n_panelists)

        else:
            full_matrix = generate_bib(n_products, k, n_panelists)

        # Convert indices to product names
        named_matrix = []
        for row in full_matrix:
            named_row = [products[i] for i in row]
            named_matrix.append(named_row)

        # Build DataFrame
        n_cols = len(named_matrix[0])
        columns = [f"Position {i+1}" for i in range(n_cols)]
        index = [f"Panelist {i+1}" for i in range(n_panelists)]
        df = pd.DataFrame(named_matrix, columns=columns, index=index)

        
        st.dataframe(df, use_container_width=True)

        if design == "Balanced Incomplete Block (BIB)":
            from itertools import combinations
            n_blocks = len(list(combinations(range(n_products), k)))
            if n_panelists > n_blocks:
                st.warning(
                    f"Your design has {n_blocks} unique blocks but {n_panelists} panelists. "
                    f"Blocks repeat after panelist {n_blocks}. "
                    f"For a fully balanced BIB, use exactly {n_blocks} panelists or a multiple of it."
                )

    with tab2:
        
        st.subheader("Blinding Codes")
        st.markdown(
            "Each product is assigned a random 3-digit code. "
            "Panelists only ever see these codes — never the real product names. "
            "Codes are locked once generated so they stay consistent across all sessions and ballots."
        )

        import random

        def generate_blinding_codes(product_names):
            # Pick unique 3-digit codes at random — no two products get the same code
            # random.sample picks n unique values from the given range
            codes = random.sample(range(100, 999), len(product_names))
            # zip pairs each product name with a code, dict() turns those pairs into a dictionary
            return {name: code for name, code in zip(product_names, codes)}

        # Show generate button only if codes haven't been locked yet
        if not st.session_state.get("codes_locked", False):
            if st.button("🎲 Generate blinding codes"):
                st.session_state["blinding_codes"] = generate_blinding_codes(config["product_names"])

        # Show the codes table if codes exist
        if st.session_state.get("blinding_codes"):
            codes = st.session_state["blinding_codes"]

            codes_df = pd.DataFrame({
                "Product (internal)": list(codes.keys()),
                "Blinding Code": list(codes.values())
            })
            st.dataframe(codes_df, use_container_width=True, hide_index=True)

            if not st.session_state.get("codes_locked", False):
                st.warning("Codes are not locked yet. You can regenerate them. Once locked they cannot change.")
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("🎲 Regenerate"):
                        st.session_state["blinding_codes"] = generate_blinding_codes(config["product_names"])
                with col2:
                    if st.button("🔒 Lock codes"):
                        st.session_state["codes_locked"] = True
            else:
                st.success("Codes are locked for this trial.")

    with tab3:
        
        st.subheader("Panelist Ballots")
        st.markdown(
            "Each panelist gets a ballot showing their serving order using blinding codes only. "
            "Generate blinding codes and lock them first before printing ballots."
        )

        if not st.session_state.get("codes_locked", False):
            st.warning("Lock your blinding codes first in the Blinding Codes tab before generating ballots.")
        elif st.session_state.get("design_matrix") is None and not st.session_state["trial_saved"]:
            st.warning("Save your configuration first.")
        else:
            codes = st.session_state["blinding_codes"]
            products = config["product_names"]
            n_products = config["num_products"]
            n_panelists = config["num_panelists"]
            design = config["design_type"]
            k = config["samples_per_panelist"]
            scale = config["scale_type"]
            attributes = config["attributes"]

            # Rebuild the design matrix the same way as tab1
            if design == "Williams Latin Square":
                matrix = generate_williams_latin_square(n_products)
                full_matrix = [matrix[i % len(matrix)] for i in range(n_panelists)]
            elif design == "Monadic Sequential":
                # Use a fixed seed so ballots match the matrix shown in tab1
                np.random.seed(42)
                full_matrix = generate_monadic_sequential(n_products, n_panelists)
            else:
                full_matrix = generate_bib(n_products, k, n_panelists)

            # Panelist selector — user picks which panelist's ballot to preview
            panelist_num = st.selectbox(
                label="Preview ballot for panelist:",
                options=[f"Panelist {i+1}" for i in range(n_panelists)]
            )
            idx = int(panelist_num.split(" ")[1]) - 1
            row = full_matrix[idx]

            st.markdown("---")
            st.markdown(f"### Ballot — {panelist_num}")
            st.markdown(f"**Trial:** {config['trial_name']}")
            st.markdown(f"**Design:** {design}")
            st.markdown("---")

            # Show each sample in serving order with blinding code
            for pos, product_idx in enumerate(row):
                product_name = products[product_idx]
                blind_code = codes[product_name]

                st.markdown(f"#### Sample {pos+1} — Code: **{blind_code}**")

                # Show rating scale appropriate to the chosen scale type
                if scale == "9-point Hedonic":
                    for attr in attributes:
                        st.slider(
                            label=f"{attr}",
                            min_value=1, max_value=9, value=5,
                            help="1 = Dislike extremely, 9 = Like extremely",
                            key=f"p{idx}_s{pos}_{attr}"
                        )

                elif scale == "100mm Visual Analogue Scale (VAS)":
                    for attr in attributes:
                        st.slider(
                            label=f"{attr}",
                            min_value=0, max_value=100, value=50,
                            help="0 = extremely negative, 100 = extremely positive",
                            key=f"p{idx}_s{pos}_{attr}"
                        )

                elif scale == "5-point Likert":
                    for attr in attributes:
                        st.select_slider(
                            label=f"{attr}",
                            options=["Strongly disagree", "Disagree", "Neutral", "Agree", "Strongly agree"],
                            key=f"p{idx}_s{pos}_{attr}"
                        )

                elif scale == "Just About Right (JAR) — 5-point":
                    for attr in attributes:
                        st.select_slider(
                            label=f"{attr}",
                            options=["Much too little", "Too little", "Just about right", "Too much", "Much too much"],
                            key=f"p{idx}_s{pos}_{attr}"
                        )

                elif scale == "Preference (forced choice: pick one)":
                    # For forced choice, only one question at the end
                    st.markdown("*(Evaluate this sample, then indicate your preference at the end)*")

                st.markdown("")

            # Forced choice preference question shown once after all samples
            if scale == "Preference (forced choice: pick one)":
                st.markdown("---")
                st.markdown("#### Which sample did you prefer?")
                blind_codes_for_row = [codes[products[i]] for i in row]
                st.radio(
                    label="Select one:",
                    options=[str(c) for c in blind_codes_for_row],
                    key=f"p{idx}_preference"
                )

            st.markdown("---")
            st.caption("Do not share this ballot with other panelists. Return to the session coordinator when complete.")

            st.markdown("---")
            st.markdown("#### 🖨️ Export ballot as PDF")

            from fpdf import FPDF

            def build_ballot_pdf(panelist_num, row, products, codes, attributes, scale, config):
                pdf = FPDF()
                pdf.add_page()

                # Title
                pdf.set_font("Helvetica", "B", 16)
                pdf.cell(0, 10, "Sensory Evaluation Ballot", ln=True, align="C")
                pdf.ln(4)

                # Trial info
                pdf.set_font("Helvetica", "", 11)
                pdf.cell(0, 8, f"Trial: {config['trial_name']}", ln=True)
                pdf.cell(0, 8, f"Panelist: {panelist_num}", ln=True)
                pdf.cell(0, 8, f"Design: {config['design_type']}", ln=True)
                pdf.ln(4)
                pdf.set_draw_color(100, 100, 100)
                pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                pdf.ln(6)

                # One section per sample
                for pos, product_idx in enumerate(row):
                    product_name = products[product_idx]
                    blind_code = codes[product_name]

                    # Sample header
                    pdf.set_font("Helvetica", "B", 13)
                    pdf.cell(0, 10, f"Sample {pos+1} - Code: {blind_code}", ln=True)
                    pdf.ln(2)

                    if scale == "9-point Hedonic":
                        for attr in attributes:
                            pdf.set_font("Helvetica", "", 11)
                            pdf.cell(0, 8, f"{attr}:", ln=True)
                            pdf.set_font("Helvetica", "I", 9)
                            pdf.cell(0, 6, "1 = Dislike extremely    2    3    4    5    6    7    8    9 = Like extremely", ln=True)
                            pdf.cell(0, 6, "Circle one:   1     2     3     4     5     6     7     8     9", ln=True)
                            pdf.ln(3)

                    elif scale == "100mm Visual Analogue Scale (VAS)":
                        for attr in attributes:
                            pdf.set_font("Helvetica", "", 11)
                            pdf.cell(0, 8, f"{attr}:", ln=True)
                            pdf.set_font("Helvetica", "I", 9)
                            pdf.cell(0, 6, "Draw a vertical line on the scale below:", ln=True)
                            # Draw the VAS line
                            y = pdf.get_y() + 4
                            pdf.line(20, y, 180, y)
                            pdf.set_font("Helvetica", "", 8)
                            pdf.set_xy(14, y + 1)
                            pdf.cell(20, 5, "0 (negative)")
                            pdf.set_xy(155, y + 1)
                            pdf.cell(30, 5, "100 (positive)")
                            pdf.ln(12)

                    elif scale == "5-point Likert":
                        for attr in attributes:
                            pdf.set_font("Helvetica", "", 11)
                            pdf.cell(0, 8, f"{attr}:", ln=True)
                            pdf.set_font("Helvetica", "I", 9)
                            pdf.cell(0, 6, "Circle one:   Strongly disagree     Disagree     Neutral     Agree     Strongly agree", ln=True)
                            pdf.ln(3)

                    elif scale == "Just About Right (JAR) — 5-point":
                        for attr in attributes:
                            pdf.set_font("Helvetica", "", 11)
                            pdf.cell(0, 8, f"{attr}:", ln=True)
                            pdf.set_font("Helvetica", "I", 9)
                            pdf.cell(0, 6, "Circle one:   Much too little     Too little     Just about right     Too much     Much too much", ln=True)
                            pdf.ln(3)

                    elif scale == "Preference (forced choice: pick one)":
                        pdf.set_font("Helvetica", "I", 10)
                        pdf.cell(0, 8, "Evaluate this sample. Indicate your preference after all samples.", ln=True)
                        pdf.ln(2)

                    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                    pdf.ln(6)

                # Forced choice final question
                if scale == "Preference (forced choice: pick one)":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(0, 10, "Which sample did you prefer?", ln=True)
                    blind_codes_for_row = [str(codes[products[i]]) for i in row]
                    pdf.set_font("Helvetica", "", 11)
                    pdf.cell(0, 8, "Circle one:   " + "          ".join(blind_codes_for_row), ln=True)
                    pdf.ln(4)

                # Footer
                pdf.set_font("Helvetica", "I", 9)
                pdf.cell(0, 8, "Do not share this ballot. Return to the session coordinator when complete.", ln=True, align="C")

                # Return as bytes
                return bytes(pdf.output())

            pdf_bytes = build_ballot_pdf(
                panelist_num, row, products, codes, attributes, scale, config
            )

            st.download_button(
                label="📄 Download PDF ballot",
                data=pdf_bytes,
                file_name=f"{config['trial_name'].replace(' ', '_')}_{panelist_num.replace(' ', '_')}_ballot.pdf",
                mime="application/pdf"
            )

            st.markdown("---")
            st.markdown("#### Download all ballots in one PDF")

            def build_all_ballots_pdf(full_matrix, products, codes, attributes, scale, config, n_panelists):
                pdf = FPDF()
                for p_idx in range(n_panelists):
                    pdf.add_page()
                    row = full_matrix[p_idx]
                    panelist_label = f"Panelist {p_idx + 1}"

                    # Title
                    pdf.set_font("Helvetica", "B", 16)
                    pdf.cell(0, 10, "Sensory Evaluation Ballot", ln=True, align="C")
                    pdf.ln(4)

                    # Trial info
                    pdf.set_font("Helvetica", "", 11)
                    pdf.cell(0, 8, f"Trial: {config['trial_name']}", ln=True)
                    pdf.cell(0, 8, f"Panelist: {panelist_label}", ln=True)
                    pdf.cell(0, 8, f"Design: {config['design_type']}", ln=True)
                    pdf.ln(4)
                    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                    pdf.ln(6)

                    for pos, product_idx in enumerate(row):
                        product_name = products[product_idx]
                        blind_code = codes[product_name]

                        pdf.set_font("Helvetica", "B", 13)
                        pdf.cell(0, 10, f"Sample {pos+1} - Code: {blind_code}", ln=True)
                        pdf.ln(2)

                        if scale == "9-point Hedonic":
                            for attr in attributes:
                                pdf.set_font("Helvetica", "", 11)
                                pdf.cell(0, 8, f"{attr}:", ln=True)
                                pdf.set_font("Helvetica", "I", 9)
                                pdf.cell(0, 6, "1 = Dislike extremely    2    3    4    5    6    7    8    9 = Like extremely", ln=True)
                                pdf.cell(0, 6, "Circle one:   1     2     3     4     5     6     7     8     9", ln=True)
                                pdf.ln(3)

                        elif scale == "100mm Visual Analogue Scale (VAS)":
                            for attr in attributes:
                                pdf.set_font("Helvetica", "", 11)
                                pdf.cell(0, 8, f"{attr}:", ln=True)
                                pdf.set_font("Helvetica", "I", 9)
                                pdf.cell(0, 6, "Draw a vertical line on the scale below:", ln=True)
                                y = pdf.get_y() + 4
                                pdf.line(20, y, 180, y)
                                pdf.set_font("Helvetica", "", 8)
                                pdf.set_xy(14, y + 1)
                                pdf.cell(20, 5, "0 (negative)")
                                pdf.set_xy(155, y + 1)
                                pdf.cell(30, 5, "100 (positive)")
                                pdf.ln(12)

                        elif scale == "5-point Likert":
                            for attr in attributes:
                                pdf.set_font("Helvetica", "", 11)
                                pdf.cell(0, 8, f"{attr}:", ln=True)
                                pdf.set_font("Helvetica", "I", 9)
                                pdf.cell(0, 6, "Circle one:   Strongly disagree     Disagree     Neutral     Agree     Strongly agree", ln=True)
                                pdf.ln(3)

                        elif scale == "Just About Right (JAR) — 5-point":
                            for attr in attributes:
                                pdf.set_font("Helvetica", "", 11)
                                pdf.cell(0, 8, f"{attr}:", ln=True)
                                pdf.set_font("Helvetica", "I", 9)
                                pdf.cell(0, 6, "Circle one:   Much too little     Too little     Just about right     Too much     Much too much", ln=True)
                                pdf.ln(3)

                        elif scale == "Preference (forced choice: pick one)":
                            pdf.set_font("Helvetica", "I", 10)
                            pdf.cell(0, 8, "Evaluate this sample. Indicate your preference after all samples.", ln=True)
                            pdf.ln(2)

                        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                        pdf.ln(6)

                    if scale == "Preference (forced choice: pick one)":
                        pdf.set_font("Helvetica", "B", 12)
                        pdf.cell(0, 10, "Which sample did you prefer?", ln=True)
                        blind_codes_for_row = [str(codes[products[i]]) for i in row]
                        pdf.set_font("Helvetica", "", 11)
                        pdf.cell(0, 8, "Circle one:   " + "          ".join(blind_codes_for_row), ln=True)
                        pdf.ln(4)

                    pdf.set_font("Helvetica", "I", 9)
                    pdf.cell(0, 8, "Do not share this ballot. Return to the session coordinator when complete.", ln=True, align="C")

                return bytes(pdf.output())

            all_ballots_pdf = build_all_ballots_pdf(
                full_matrix, products, codes, attributes, scale, config, n_panelists
            )

            st.download_button(
                label="📄 Download all ballots (one PDF)",
                data=all_ballots_pdf,
                file_name=f"{config['trial_name'].replace(' ', '_')}_all_ballots.pdf",
                mime="application/pdf"
            )

    with tab4:
        
        st.subheader("Power Calculator")
        st.markdown(
            "Estimates the minimum panel size needed to detect a real difference between products. "
            "Based on your chosen effect size, significance level, and desired statistical power."
        )

        from scipy import stats

        st.markdown("---")
        st.markdown("#### What do these terms mean?")

        with st.expander("📖 Quick reference — effect size, alpha, power"):
            st.markdown(
                "**Effect size (d)** — how large a difference you expect between products. "
                "In sensory science, d=0.5 is a medium difference (noticeable but not dramatic), "
                "d=0.8 is large (clearly different products), d=0.2 is small (subtle difference)."
            )
            st.markdown(
                "**Significance level (alpha)** — how much risk you accept of saying products differ "
                "when they actually don't. Standard is 0.05 (5% risk)."
            )
            st.markdown(
                "**Power (1 - beta)** — probability of detecting a real difference if one exists. "
                "0.80 means you have an 80% chance of catching a real effect. Standard minimum."
            )

        st.markdown("---")
        st.markdown("#### Inputs")

        col1, col2, col3 = st.columns(3)

        with col1:
            effect_size = st.slider(
                label="Expected effect size (d)",
                min_value=0.1, max_value=2.0, value=0.5, step=0.1,
                help="0.2 = small, 0.5 = medium, 0.8 = large"
            )

        with col2:
            alpha = st.selectbox(
                label="Significance level (alpha)",
                options=[0.01, 0.05, 0.10],
                index=1,
                help="Risk of false positive. 0.05 is the standard."
            )

        with col3:
            power = st.selectbox(
                label="Desired power (1 - beta)",
                options=[0.70, 0.80, 0.90, 0.95],
                index=1,
                help="0.80 is the standard minimum for sensory studies."
            )

        # Power calculation using a two-tailed t-test approximation
        # This is the standard formula for comparing two means:
        #   n = 2 * ((z_alpha/2 + z_beta) / d)^2
        # where z_alpha/2 and z_beta are the z-scores for alpha and power
        z_alpha = stats.norm.ppf(1 - alpha / 2)   # two-tailed
        z_beta = stats.norm.ppf(power)
        n_required = int(np.ceil(2 * ((z_alpha + z_beta) / effect_size) ** 2))

        st.markdown("---")
        st.markdown("#### Result")

        col_a, col_b, col_c = st.columns(3)
        with col_a:
            st.metric(label="Minimum panelists required", value=n_required)
        with col_b:
            st.metric(label="Your current panel size", value=config["num_panelists"])
        with col_c:
            difference = config["num_panelists"] - n_required
            st.metric(
                label="Surplus / deficit",
                value=difference,
                delta=difference,
                # Green if surplus, red if deficit
            )

        if config["num_panelists"] < n_required:
            st.error(
                f"Your panel of {config['num_panelists']} is underpowered for these parameters. "
                f"You need at least {n_required} panelists to detect a d={effect_size} effect "
                f"at alpha={alpha} with {int(power*100)}% power."
            )
        else:
            st.success(
                f"Your panel of {config['num_panelists']} is sufficient. "
                f"You can detect a d={effect_size} effect at alpha={alpha} with {int(power*100)}% power."
            )

        # Show a table of required n across a range of effect sizes
        st.markdown("---")
        st.markdown("#### Reference table — minimum panelists at different effect sizes")
        st.caption(f"Alpha={alpha}, Power={int(power*100)}%")

        effect_sizes = [0.2, 0.3, 0.5, 0.8, 1.0, 1.5, 2.0]
        n_values = []
        for d in effect_sizes:
            n = int(np.ceil(2 * ((z_alpha + z_beta) / d) ** 2))
            n_values.append(n)

        ref_df = pd.DataFrame({
            "Effect size (d)": effect_sizes,
            "Interpretation": ["Very small", "Small", "Medium", "Large", "Very large", "Huge", "Extreme"],
            "Min panelists required": n_values
        })
        st.dataframe(ref_df, use_container_width=True, hide_index=True)

    with tab5:
        
        st.subheader("Data Entry Template")
        st.markdown(
            "Download a pre-formatted Excel file ready for data entry after your trial. "
            "One row per panelist per sample, with columns for each attribute. "
            "Import directly into R or Python for analysis."
        )

        if not st.session_state.get("codes_locked", False):
            st.warning("Lock your blinding codes first in the Blinding Codes tab.")
        else:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            codes = st.session_state["blinding_codes"]
            products = config["product_names"]
            n_products = config["num_products"]
            n_panelists = config["num_panelists"]
            design = config["design_type"]
            k = config["samples_per_panelist"]
            scale = config["scale_type"]
            attributes = config["attributes"]

            # Rebuild design matrix
            if design == "Williams Latin Square":
                matrix = generate_williams_latin_square(n_products)
                full_matrix = [matrix[i % len(matrix)] for i in range(n_panelists)]
            elif design == "Monadic Sequential":
                np.random.seed(42)
                full_matrix = generate_monadic_sequential(n_products, n_panelists)
            else:
                full_matrix = generate_bib(n_products, k, n_panelists)

            def build_excel(full_matrix, products, codes, attributes, scale, config):
                wb = Workbook()
                ws = wb.active
                ws.title = "Data Entry"

                # Header style — dark background, white bold text
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="2F4F4F")
                center = Alignment(horizontal="center")

                # Build column headers
                headers = ["Panelist", "Sample Position", "Blinding Code", "Product (internal)"]
                if scale == "Preference (forced choice: pick one)":
                    headers.append("Preference (enter winning code)")
                else:
                    headers += attributes

                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center

                # Fill rows — one row per panelist per sample
                row_num = 2
                for p_idx, product_row in enumerate(full_matrix):
                    for pos, product_idx in enumerate(product_row):
                        product_name = products[product_idx]
                        blind_code = codes[product_name]

                        ws.cell(row=row_num, column=1, value=f"Panelist {p_idx+1}")
                        ws.cell(row=row_num, column=2, value=pos + 1)
                        ws.cell(row=row_num, column=3, value=blind_code)
                        ws.cell(row=row_num, column=4, value=product_name)

                        # Leave attribute columns blank for data entry
                        for col in range(5, len(headers) + 1):
                            ws.cell(row=row_num, column=col, value="")

                        row_num += 1

                # Auto-size columns
                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max_len + 4

                # Save to a bytes buffer — this is what Streamlit will download
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer

            # Preview the structure as a table before downloading
            st.markdown("**Preview — first 5 rows**")
            preview_rows = []
            for p_idx, product_row in enumerate(full_matrix[:3]):
                for pos, product_idx in enumerate(product_row):
                    product_name = products[product_idx]
                    blind_code = codes[product_name]
                    row = {
                        "Panelist": f"Panelist {p_idx+1}",
                        "Position": pos + 1,
                        "Blinding Code": blind_code,
                        "Product (internal)": product_name,
                    }
                    if scale != "Preference (forced choice: pick one)":
                        for attr in attributes:
                            row[attr] = ""
                    else:
                        row["Preference"] = ""
                    preview_rows.append(row)

            preview_df = pd.DataFrame(preview_rows[:5])
            st.dataframe(preview_df, use_container_width=True, hide_index=True)

            st.markdown("")

            # Build and offer the Excel file for download
            excel_buffer = build_excel(full_matrix, products, codes, attributes, scale, config)

            st.download_button(
                label="📥 Download Excel data entry template",
                data=excel_buffer,
                file_name=f"{config['trial_name'].replace(' ', '_')}_data_entry.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.caption(
                "The Product (internal) column is for your reference only — remove it before "
                "sharing with panelists or data entry staff. Attribute columns are blank and ready for entry."
            )
    with tab6:
        st.subheader("Share Links")
        st.markdown(
            "Publish this trial to the database and generate a unique ballot link for each panelist. "
            "Each link opens a personalized digital ballot on any device."
        )

        if not st.session_state.get("codes_locked", False):
            st.warning("Lock your blinding codes first in the Blinding Codes tab.")
        else:
            codes = st.session_state["blinding_codes"]
            products = config["product_names"]
            n_products = config["num_products"]
            n_panelists = config["num_panelists"]
            design = config["design_type"]
            k = config["samples_per_panelist"]

            # Rebuild design matrix
            if design == "Williams Latin Square":
                matrix = generate_williams_latin_square(n_products)
                full_matrix = [matrix[i % len(matrix)] for i in range(n_panelists)]
            elif design == "Monadic Sequential":
                np.random.seed(42)
                full_matrix = generate_monadic_sequential(n_products, n_panelists)
            else:
                full_matrix = generate_bib(n_products, k, n_panelists)

            # Publish button — saves trial and panelists to Supabase
            if not st.session_state.get("trial_published", False):
                st.info("Click below to publish this trial and generate panelist links.")
                if st.button("🚀 Publish trial and generate links", type="primary"):
                    try:
                        # Save trial config to Supabase
                        trial_result = supabase.table("trials").insert({
                            "trial_name": config["trial_name"],
                            "config": config
                        }).execute()

                        trial_id = trial_result.data[0]["id"]

                        # Generate one panelist record per panelist
                        panelist_tokens = []
                        for p_idx in range(n_panelists):
                            row = full_matrix[p_idx]

                            # Build serving order as list of dicts
                            # Each dict has the product name and blinding code
                            serving_order = [
                                {
                                    "product_name": products[i],
                                    "blinding_code": codes[products[i]]
                                }
                                for i in row
                            ]

                            # Generate a unique random token for this panelist
                            # secrets.token_urlsafe generates a cryptographically secure random string
                            token = secrets.token_urlsafe(16)

                            supabase.table("panelists").insert({
                                "trial_id": trial_id,
                                "panelist_number": p_idx + 1,
                                "access_token": token,
                                "serving_order": serving_order,
                                "submitted": False
                            }).execute()

                            panelist_tokens.append(token)

                        # Save to session state so links persist
                        st.session_state["trial_published"] = True
                        st.session_state["trial_id"] = trial_id
                        st.session_state["panelist_tokens"] = panelist_tokens
                        st.rerun()

                    except Exception as e:
                        st.error(f"Publishing failed: {e}")

            else:
                trial_id = st.session_state["trial_id"]
                tokens = st.session_state["panelist_tokens"]

                st.success(f"Trial published. {len(tokens)} panelist links generated.")
                st.markdown("---")

                # Get the app's base URL
                # On Streamlit Cloud this will be the public URL
                # Locally it will be localhost
                base_url = st.session_state.get("base_url", "http://localhost:8501")
                custom_url = st.text_input(
                    label="App base URL",
                    value=base_url,
                    help="Change this to your Streamlit Cloud URL after deployment."
                )
                st.session_state["base_url"] = custom_url

                st.markdown("---")
                st.markdown("#### Panelist links")
                st.caption("Share each link with the corresponding panelist. Each link works only once.")

                # Display links as a table
                import pandas as pd
                links_data = []
                for i, token in enumerate(tokens):
                    link = f"{custom_url}/ballot?token={token}"
                    links_data.append({
                        "Panelist": f"Panelist {i+1}",
                        "Link": link
                    })

                links_df = pd.DataFrame(links_data)
                st.dataframe(links_df, use_container_width=True, hide_index=True)

                # Download links as Excel
                import io
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Panelist Links"
                ws.append(["Panelist", "Link"])
                for row in links_data:
                    ws.append([row["Panelist"], row["Link"]])

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                st.download_button(
                    label="📥 Download all links as Excel",
                    data=buf,
                    file_name=f"{config['trial_name'].replace(' ', '_')}_links.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.markdown("---")
                st.markdown("#### Response status")

                if st.button("🔄 Refresh status"):
                    st.rerun()

                # Check how many panelists have submitted
                status_result = supabase.table("panelists").select("panelist_number, submitted").eq("trial_id", trial_id).execute()
                status_df = pd.DataFrame(status_result.data)
                status_df.columns = ["Panelist Number", "Submitted"]
                status_df["Submitted"] = status_df["Submitted"].map({True: "Yes", False: "No"})
                st.dataframe(status_df, use_container_width=True, hide_index=True)

                submitted_count = sum(1 for r in status_result.data if r["submitted"])
                st.markdown("---")
                st.markdown("#### Download results")

                if submitted_count == 0:
                    st.info("No responses yet.")
                else:
                    if st.button("📥 Download responses as Excel"):
                        # Pull all responses for this trial from Supabase
                        responses_result = supabase.table("responses").select("*").eq("trial_id", trial_id).execute()
                        responses = responses_result.data

                        # Pull panelist info for labeling
                        panelists_result = supabase.table("panelists").select("id, panelist_number").eq("trial_id", trial_id).execute()
                        panelist_map = {p["id"]: p["panelist_number"] for p in panelists_result.data}

                        # Build rows for Excel
                        rows = []
                        for r in responses:
                            base_row = {
                                "Panelist": f"Panelist {panelist_map.get(r['panelist_id'], '?')}",
                                "Sample Position": r["sample_position"],
                                "Blinding Code": r["blinding_code"],
                            }
                            # ratings is a dict of attribute -> value
                            # flatten it into columns
                            for attr, val in r["ratings"].items():
                                base_row[attr] = val
                            rows.append(base_row)

                        results_df = pd.DataFrame(rows)
                        results_df = results_df.sort_values(["Panelist", "Sample Position"])

                        # Write to Excel
                        import io
                        from openpyxl import Workbook
                        buf = io.BytesIO()
                        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                            results_df.to_excel(writer, index=False, sheet_name="Responses")
                        buf.seek(0)

                        st.download_button(
                            label="📥 Download Excel",
                            data=buf,
                            file_name=f"{config['trial_name'].replace(' ', '_')}_responses.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        # Preview
                        st.dataframe(results_df, use_container_width=True, hide_index=True)
                